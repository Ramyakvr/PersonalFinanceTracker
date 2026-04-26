"""``python manage.py inspect_broker <path>``

Dump a broker export's structure so new adapters can be mapped quickly. Handles:

* ``.csv`` — prints header + first 5 rows.
* ``.xlsx`` — finds the likely header row (first row with text cells in ≥ 3
  consecutive columns), prints it + first 5 data rows.
* ``.pdf`` — prints the first 30 lines of extracted text (requires
  ``pdfplumber``; if not installed, falls back to ``pypdf``).

This is a throwaway developer tool, not a user-facing feature. Use it when a
broker ships a format we don't yet have an adapter for (e.g. Aionion) to
decide which columns map to which ``NormalizedTrade`` fields.
"""

from __future__ import annotations

import csv
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

MAX_SAMPLE_ROWS = 5


class Command(BaseCommand):
    help = "Inspect a broker export file's structure (csv | xlsx | pdf)."

    def add_arguments(self, parser):
        parser.add_argument("path", type=str, help="Path to the broker export file")

    def handle(self, *args, path: str, **options):
        p = Path(path).expanduser()
        if not p.exists():
            raise CommandError(f"File not found: {p}")
        ext = p.suffix.lower()
        self.stdout.write(self.style.MIGRATE_HEADING(f"Inspecting: {p}"))
        self.stdout.write(f"  size: {p.stat().st_size} bytes")
        self.stdout.write(f"  extension: {ext}")

        if ext == ".csv":
            self._inspect_csv(p)
        elif ext == ".xlsx":
            self._inspect_xlsx(p)
        elif ext == ".pdf":
            self._inspect_pdf(p)
        else:
            raise CommandError(f"Unsupported extension: {ext}")

    # -- CSV ----------------------------------------------------------------

    def _inspect_csv(self, p: Path) -> None:
        with p.open(newline="", encoding="utf-8-sig") as fh:
            reader = csv.reader(fh)
            rows = [row for i, row in enumerate(reader) if i < MAX_SAMPLE_ROWS + 1]
        if not rows:
            self.stdout.write(self.style.WARNING("Empty CSV."))
            return
        self.stdout.write(self.style.SUCCESS("Header:"))
        self.stdout.write(f"  {rows[0]}")
        self.stdout.write(self.style.SUCCESS("Rows:"))
        for r in rows[1:]:
            self.stdout.write(f"  {r}")

    # -- XLSX ---------------------------------------------------------------

    def _inspect_xlsx(self, p: Path) -> None:
        from openpyxl import load_workbook

        wb = load_workbook(p, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            self.stdout.write(self.style.SUCCESS(f"Sheet: {sheet_name}"))
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
            header_idx = self._guess_header_row(rows)
            if header_idx < 0:
                self.stdout.write(self.style.WARNING("  No plausible header row found."))
                continue
            self.stdout.write(f"  Header row (index {header_idx}):")
            self.stdout.write(f"  {rows[header_idx]}")
            self.stdout.write(f"  First {MAX_SAMPLE_ROWS} data rows:")
            for row in rows[header_idx + 1 : header_idx + 1 + MAX_SAMPLE_ROWS]:
                self.stdout.write(f"  {row}")
        wb.close()

    def _guess_header_row(self, rows: list[list]) -> int:
        """Header = the first row with >= 3 consecutive non-null text cells."""
        for i, row in enumerate(rows):
            text_run = 0
            for cell in row or ():
                if cell is not None and isinstance(cell, str) and cell.strip():
                    text_run += 1
                    if text_run >= 3:
                        return i
                else:
                    text_run = 0
        return -1

    # -- PDF ----------------------------------------------------------------

    def _inspect_pdf(self, p: Path) -> None:
        try:
            import pdfplumber  # noqa: F401
        except ImportError:
            self.stdout.write(
                self.style.WARNING(
                    "pdfplumber not installed — falling back to raw text extraction."
                )
            )
            self._pdf_via_pypdf(p)
            return
        import pdfplumber

        with pdfplumber.open(p) as pdf:
            self.stdout.write(self.style.SUCCESS(f"Pages: {len(pdf.pages)}"))
            for i, page in enumerate(pdf.pages[:2], start=1):
                self.stdout.write(f"--- Page {i} text (first 30 lines) ---")
                text = page.extract_text() or ""
                for line in text.splitlines()[:30]:
                    self.stdout.write(f"  {line}")
                self.stdout.write(f"--- Page {i} tables ---")
                tables = page.extract_tables() or []
                for t_idx, table in enumerate(tables):
                    self.stdout.write(f"  Table {t_idx} ({len(table)} rows):")
                    for row in table[:3]:
                        self.stdout.write(f"    {row}")

    def _pdf_via_pypdf(self, p: Path) -> None:
        try:
            from pypdf import PdfReader
        except ImportError as exc:
            raise CommandError(
                "Install pdfplumber (preferred) or pypdf to inspect PDFs."
            ) from exc
        r = PdfReader(str(p))
        self.stdout.write(self.style.SUCCESS(f"Pages: {len(r.pages)}"))
        for i, page in enumerate(r.pages[:2], start=1):
            self.stdout.write(f"--- Page {i} text (first 30 lines) ---")
            text = page.extract_text() or ""
            for line in text.splitlines()[:30]:
                self.stdout.write(f"  {line}")
