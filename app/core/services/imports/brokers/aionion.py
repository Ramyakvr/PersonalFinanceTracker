"""Aionion (Aionion Capital) XLSX parser.

Two file types from the broker portal:

* ``Equity_Trades_<period>.xlsx`` -- single sheet ``Nexus Statement``.
  Preamble rows 0-10 carry client metadata (``CLIENT ID``, ``NAME``,
  ``PAN``, totals). Header at the row whose column A is ``SYMBOL``.
  Data columns: ``SYMBOL, ISIN, DATE, PRODUCT, TYPE, QUANTITY, PRICE,
  NET VALUE``. Date format is ``DD/MM/YYYY``. ``TYPE`` is
  ``BUY``/``SELL``. A trailing ``TOTALS`` row terminates the data.

* ``Dividend_Income_Audit_<client>_<period>.xlsx`` -- two sheets:
    - ``Summary`` -- one row per stock with ``ASSET CLASS``, ``ISIN``,
      ``QUANTITY``, ``RECENT EX-DATE``, ``RECENT AMOUNT``,
      ``TOTAL INCOME``. Used to build the ``stock name -> ISIN`` map
      since the detailed sheet doesn't repeat ISIN per row.
    - ``Detailed Audit`` -- per-stock event lists. Each stock's section
      starts with ``STOCK: <NAME>``, then the ``RECORD DATE | QUANTITY |
      DIVIDEND PER UNIT | TOTAL AMOUNT`` sub-header, then one row per
      dividend event, then a ``-,-,-,-`` separator row.

Aionion does not provide a broker-native trade ID; ``trade_ref`` is
synthesized as ``aionion:{date}:{isin}:{type}:{qty}:{price}:{occurrence}``
where ``occurrence`` is a per-key counter so re-imports stay idempotent
even if rows are reordered.

Charges (brokerage / STT / GST / etc.) are not broken out in either
file; we default them to zero.

Corporate actions are not exported by Aionion in any machine-readable
form -- ``parse_corporate_actions`` returns an empty iterator.
"""

from __future__ import annotations

from collections.abc import Iterable
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook

from core.services.imports.brokers.base import (
    BrokerFormatError,
    NormalizedCA,
    NormalizedDividend,
    NormalizedTrade,
)

ZERO = Decimal(0)

TRADEBOOK_REQUIRED = {
    "SYMBOL",
    "ISIN",
    "DATE",
    "TYPE",
    "QUANTITY",
    "PRICE",
}

DIV_SUMMARY_REQUIRED = {
    "ASSET CLASS",
    "ISIN",
}
DIV_DETAIL_SUBHEADER = {
    "RECORD DATE",
    "QUANTITY",
    "DIVIDEND PER UNIT",
    "TOTAL AMOUNT",
}

TERMINAL_ROW_MARKER = "totals"
STOCK_HEADER_PREFIX = "STOCK:"
DETAIL_SEPARATOR_CELL = "-"


def _load_workbook(file_bytes: bytes):
    try:
        return load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # openpyxl raises various types
        raise BrokerFormatError(f"Not a valid XLSX: {exc}") from exc


def _strip(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _to_decimal(value) -> Decimal:
    s = _strip(value).replace(",", "")
    if not s or s == DETAIL_SEPARATOR_CELL:
        return ZERO
    try:
        return Decimal(s)
    except InvalidOperation as exc:
        raise BrokerFormatError(f"Not a decimal: {value!r}") from exc


def _parse_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = _strip(value)
    try:
        return datetime.strptime(s, "%d/%m/%Y").date()
    except ValueError as exc:
        raise BrokerFormatError(f"Unrecognized date: {value!r}") from exc


def _find_header(rows: list, required: set[str]) -> int:
    """Return the index of the first row whose stripped string cells are a
    superset of ``required``. ``BrokerFormatError`` if not found."""
    for idx, row in enumerate(rows):
        if not row:
            continue
        cells = {_strip(c).upper() for c in row if c is not None}
        if required.issubset(cells):
            return idx
    raise BrokerFormatError(f"Header row not found. Required: {sorted(required)}")


def _is_totals_row(row) -> bool:
    if not row:
        return False
    first = _strip(row[0]).lower()
    return first == TERMINAL_ROW_MARKER


class AionionAdapter:
    key = "aionion"
    display_name = "Aionion"
    tradebook_extensions: tuple[str, ...] = (".xlsx",)
    dividend_extensions: tuple[str, ...] = (".xlsx",)

    # -- tradebook ----------------------------------------------------------

    def parse_tradebook(
        self, file_bytes: bytes, *, account_label: str = "Main"
    ) -> Iterable[NormalizedTrade]:
        wb = _load_workbook(file_bytes)
        try:
            ws = wb.active
            rows = [list(r) for r in ws.iter_rows(values_only=True)]
        finally:
            wb.close()

        hdr_idx = _find_header(rows, TRADEBOOK_REQUIRED)
        header = [_strip(c).upper() for c in rows[hdr_idx]]
        col = {name: i for i, name in enumerate(header) if name}

        occurrence_count: dict[tuple, int] = {}
        for raw_row in rows[hdr_idx + 1 :]:
            if _is_totals_row(raw_row):
                break
            if not raw_row:
                continue
            symbol = _strip(raw_row[col["SYMBOL"]] if col.get("SYMBOL") is not None else None)
            if not symbol:
                continue
            try:
                trade_dt = _parse_date(raw_row[col["DATE"]])
                quantity = _to_decimal(raw_row[col["QUANTITY"]])
                price = _to_decimal(raw_row[col["PRICE"]])
            except BrokerFormatError:
                continue
            if quantity <= ZERO or price <= ZERO:
                continue
            side = _strip(raw_row[col["TYPE"]]).upper()
            if side not in ("BUY", "SELL"):
                continue
            isin = _strip(raw_row[col["ISIN"]])
            product = _strip(raw_row[col["PRODUCT"]]) if col.get("PRODUCT") is not None else ""

            bucket = (trade_dt, isin, side, quantity, price)
            occurrence_count[bucket] = occurrence_count.get(bucket, 0) + 1
            occurrence = occurrence_count[bucket]
            trade_ref = (
                f"aionion:{trade_dt.isoformat()}:{isin}:{side}:{quantity}:{price}:{occurrence}"
            )
            yield NormalizedTrade(
                broker_key=self.key,
                account_label=account_label,
                trade_ref=trade_ref,
                trade_date=trade_dt,
                isin=isin,
                symbol=symbol,
                name=symbol,
                side=side,
                quantity=quantity,
                price=price,
                currency="INR",
                raw={"product": product, "source": "aionion_trades_xlsx"},
            )

    # -- dividends ----------------------------------------------------------

    def parse_dividends(
        self, file_bytes: bytes, *, account_label: str = "Main"
    ) -> Iterable[NormalizedDividend]:
        wb = _load_workbook(file_bytes)
        try:
            name_to_isin = self._build_summary_isin_map(wb)
            yield from self._iter_detail_dividends(wb, name_to_isin, account_label)
        finally:
            wb.close()

    def _build_summary_isin_map(self, wb) -> dict[str, str]:
        """Map ``ASSET CLASS`` (uppercased, stripped) to ``ISIN`` from the
        Summary sheet. Raise ``BrokerFormatError`` if the sheet is missing
        or malformed -- the Detailed Audit sheet alone cannot be parsed.
        """
        if "Summary" not in wb.sheetnames:
            raise BrokerFormatError("Aionion dividend file missing 'Summary' sheet.")
        ws = wb["Summary"]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        hdr_idx = _find_header(rows, DIV_SUMMARY_REQUIRED)
        header = [_strip(c).upper() for c in rows[hdr_idx]]
        col = {name: i for i, name in enumerate(header) if name}
        mapping: dict[str, str] = {}
        for raw_row in rows[hdr_idx + 1 :]:
            if _is_totals_row(raw_row):
                break
            if not raw_row:
                continue
            name = _strip(raw_row[col["ASSET CLASS"]]).upper()
            isin = _strip(raw_row[col["ISIN"]])
            if name and isin:
                mapping[name] = isin
        return mapping

    def _iter_detail_dividends(
        self, wb, name_to_isin: dict[str, str], account_label: str
    ) -> Iterable[NormalizedDividend]:
        if "Detailed Audit" not in wb.sheetnames:
            raise BrokerFormatError("Aionion dividend file missing 'Detailed Audit' sheet.")
        ws = wb["Detailed Audit"]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]

        current_stock: str = ""
        current_isin: str = ""
        in_data_block = False

        for raw_row in rows:
            if not raw_row:
                continue
            first = _strip(raw_row[0])
            first_upper = first.upper()

            if first_upper.startswith(STOCK_HEADER_PREFIX):
                # ``STOCK: <NAME>`` -- start of a new section
                current_stock = first[len(STOCK_HEADER_PREFIX) :].strip()
                current_isin = name_to_isin.get(current_stock.upper(), "")
                in_data_block = False
                continue

            if first_upper == "RECORD DATE":
                # Sub-header signals data rows follow until separator
                in_data_block = True
                continue

            if first == DETAIL_SEPARATOR_CELL or first_upper == "":
                in_data_block = False
                continue

            if not in_data_block or not current_isin:
                continue

            try:
                ex_dt = _parse_date(raw_row[0])
                qty = _to_decimal(raw_row[1])
                amount = _to_decimal(raw_row[3])
            except (BrokerFormatError, IndexError):
                continue
            if amount <= ZERO:
                continue

            yield NormalizedDividend(
                broker_key=self.key,
                account_label=account_label,
                isin=current_isin,
                symbol=current_stock,
                ex_date=ex_dt,
                pay_date=None,
                amount_gross=amount,
                amount_net=amount,
                tds=ZERO,
                dividend_per_share=None,
                quantity=qty if qty > ZERO else None,
                currency="INR",
                raw={"source": "aionion_dividends_xlsx"},
            )

    # -- corporate actions --------------------------------------------------

    def parse_corporate_actions(
        self, file_bytes: bytes, *, account_label: str = "Main"
    ) -> Iterable[NormalizedCA]:
        # Aionion does not export corporate actions in a structured form.
        return iter(())

    # -- client id ---------------------------------------------------------

    def parse_client_id(self, file_bytes: bytes) -> str:
        """Pull ``CLIENT ID`` out of the preamble (column A == 'CLIENT ID',
        column B == the value). Works for both the equity-trades workbook
        and the dividend audit workbook (Summary + Detailed Audit sheets
        all carry it). Returns ``""`` if not found or the bytes are not a
        valid workbook (the importer already error-reports such files via
        its own parsers; a blank client id keeps detection best-effort).
        """
        try:
            wb = _load_workbook(file_bytes)
        except BrokerFormatError:
            return ""
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if row_idx > 12:
                        break
                    for i, cell in enumerate(row):
                        if _strip(cell).upper() == "CLIENT ID":
                            for nxt in row[i + 1 :]:
                                val = _strip(nxt)
                                if val:
                                    return val.upper()
        finally:
            wb.close()
        return ""
