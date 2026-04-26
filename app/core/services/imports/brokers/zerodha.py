"""Zerodha (Kite) XLSX parser.

Handles two file types from Zerodha Console:

* ``tradebook-<client>-EQ.xlsx``: row-per-trade equity tradebook. Header row
  starts at the first row whose column A is ``"Symbol"`` (rows 0-12 are
  the broker/client preamble). Columns: ``Symbol, ISIN, Trade Date,
  Exchange, Segment, Series, Trade Type, Auction, Quantity, Price,
  Trade ID, Order ID, Order Execution Time``.
* ``dividends-<client>-<fy>.xlsx``: equity dividends. Columns:
  ``Symbol, ISIN, Ex-Date, Quantity, Dividend Per Share, Net Dividend
  Amount``. A trailing ``"Total Dividend Amount"`` summary row is skipped.

Zerodha tradebooks do **not** include charges (brokerage/STT/stamp/etc.) --
those live in a separate P&L / charges CSV. We default all charge fields
to zero; downstream UI flags instruments with ``total_charges == 0``
pending a charges import.

Dividend files give only **Ex-Date** (no pay-date); Zerodha's own
footnote states dividends credit within 30-45 days. The normalized
record leaves ``pay_date=None``; the XIRR builder falls back to
``ex_date + 35 days``.
"""

from __future__ import annotations

from collections.abc import Iterable
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from core.services.imports.brokers.base import (
    BrokerFormatError,
    NormalizedCA,
    NormalizedDividend,
    NormalizedTrade,
)

IST = ZoneInfo("Asia/Kolkata")
ZERO = Decimal(0)

TRADEBOOK_REQUIRED = {
    "Symbol",
    "ISIN",
    "Trade Date",
    "Exchange",
    "Segment",
    "Trade Type",
    "Quantity",
    "Price",
    "Trade ID",
}
DIVIDEND_REQUIRED = {
    "Symbol",
    "ISIN",
    "Ex-Date",
    "Quantity",
    "Dividend Per Share",
    "Net Dividend Amount",
}
TERMINAL_DIV_ROW = "total dividend amount"


def _load_rows(file_bytes: bytes) -> list[list]:
    """Return the first sheet's rows as lists of cell values.

    We do **not** use ``read_only=True`` -- Zerodha XLSX files ship with
    stale sheet dimensions that cause the streaming reader to report only
    a single ``[None]`` row. Tradebooks are small (typically < 1000 rows)
    so full-mode parsing is cheap.
    """
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:  # openpyxl raises varying types depending on cause
        raise BrokerFormatError(f"Not a valid XLSX: {exc}") from exc
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _find_header(
    rows: list[list], required: set[str]
) -> tuple[int, list[str]]:
    for idx, row in enumerate(rows):
        if not row:
            continue
        header = [str(c).strip() if c is not None else "" for c in row]
        if required.issubset(set(header)):
            return idx, header
    raise BrokerFormatError(
        f"Header row not found. Required columns: {sorted(required)}"
    )


def _to_decimal(value, default: Decimal = ZERO) -> Decimal:
    if value is None or value == "":
        return default
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError) as exc:
        raise BrokerFormatError(f"Not a decimal: {value!r}") from exc


def _to_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%b-%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise BrokerFormatError(f"Unrecognized date: {value!r}")


def _to_datetime(value) -> datetime | None:
    """Parse a Zerodha ``Order Execution Time`` into an IST-aware datetime.

    Zerodha timestamps are local Exchange time (IST). We localise them
    explicitly so they round-trip cleanly through Django's ``USE_TZ=True``
    ORM layer.
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=IST)
    s = str(value).strip()
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            parsed = datetime.strptime(s, fmt)
            return parsed.replace(tzinfo=IST)
        except ValueError:
            continue
    return None


EQUITY_SEGMENT = "EQ"
MF_SEGMENT = "MF"
TRADEABLE_SEGMENTS = (EQUITY_SEGMENT, MF_SEGMENT)


def _segment(row: dict) -> str:
    return str(row.get("Segment", "") or "").strip().upper()


def _is_tradeable_segment(row: dict) -> bool:
    """Accept equity (EQ) + mutual fund (MF) rows. Skip F&O, CDS, COM, etc."""
    return _segment(row) in TRADEABLE_SEGMENTS


def _kind_for_segment(row: dict) -> str:
    """Map Zerodha's Segment column to ``Instrument.kind`` values."""
    return "MF" if _segment(row) == MF_SEGMENT else "STOCK"


def _is_auction(row: dict) -> bool:
    raw = row.get("Auction")
    if raw is None or raw == "":
        return False
    try:
        return int(raw) != 0
    except (TypeError, ValueError):
        return str(raw).strip().lower() in ("true", "yes", "y", "t")


class ZerodhaAdapter:
    key = "zerodha"
    display_name = "Zerodha (Kite)"
    tradebook_extensions: tuple[str, ...] = (".xlsx",)
    dividend_extensions: tuple[str, ...] = (".xlsx",)

    def parse_tradebook(
        self, file_bytes: bytes, *, account_label: str = "Main"
    ) -> Iterable[NormalizedTrade]:
        rows = _load_rows(file_bytes)
        hdr_idx, header = _find_header(rows, TRADEBOOK_REQUIRED)
        col = {name: i for i, name in enumerate(header)}
        for raw_row in rows[hdr_idx + 1 :]:
            if not raw_row:
                continue
            row_dict = {
                name: raw_row[i] if i < len(raw_row) else None
                for name, i in col.items()
            }
            symbol = str(row_dict.get("Symbol") or "").strip()
            if not symbol:
                continue
            if not _is_tradeable_segment(row_dict):
                continue
            if _is_auction(row_dict):
                continue
            kind = _kind_for_segment(row_dict)
            side = str(row_dict.get("Trade Type") or "").strip().upper()
            if side not in ("BUY", "SELL"):
                continue
            try:
                quantity = _to_decimal(row_dict.get("Quantity"))
                price = _to_decimal(row_dict.get("Price"))
                trade_dt = _to_date(row_dict.get("Trade Date"))
            except BrokerFormatError:
                continue
            if quantity <= ZERO or price <= ZERO:
                continue
            trade_id = str(row_dict.get("Trade ID") or "").strip()
            order_id = str(row_dict.get("Order ID") or "").strip()
            if not trade_id:
                continue
            trade_ref = f"zerodha:{trade_id}"
            yield NormalizedTrade(
                broker_key=self.key,
                account_label=account_label,
                trade_ref=trade_ref,
                trade_date=trade_dt,
                exec_time=_to_datetime(row_dict.get("Order Execution Time")),
                isin=str(row_dict.get("ISIN") or "").strip(),
                # Zerodha uses the full fund name in the Symbol column for MF
                # tradebooks -- it often exceeds the 40-char Instrument.exchange_symbol
                # field. ``_get_or_create_instrument`` truncates long symbols to "".
                symbol=symbol,
                name=symbol,
                side=side,
                quantity=quantity,
                price=price,
                instrument_kind=kind,
                currency="INR",
                exchange=str(row_dict.get("Exchange") or "").strip(),
                raw={
                    "trade_id": trade_id,
                    "order_id": order_id,
                    "segment": row_dict.get("Segment"),
                    "series": row_dict.get("Series"),
                },
            )

    def parse_dividends(
        self, file_bytes: bytes, *, account_label: str = "Main"
    ) -> Iterable[NormalizedDividend]:
        rows = _load_rows(file_bytes)
        hdr_idx, header = _find_header(rows, DIVIDEND_REQUIRED)
        col = {name: i for i, name in enumerate(header)}
        for raw_row in rows[hdr_idx + 1 :]:
            if not raw_row:
                continue
            symbol_raw = raw_row[col["Symbol"]] if col["Symbol"] < len(raw_row) else None
            if symbol_raw is None:
                continue
            symbol = str(symbol_raw).strip()
            if not symbol:
                continue
            if symbol.lower().startswith(TERMINAL_DIV_ROW):
                break  # end-of-data marker
            row_dict = {
                name: raw_row[i] if i < len(raw_row) else None
                for name, i in col.items()
            }
            try:
                ex_date = _to_date(row_dict.get("Ex-Date"))
                net = _to_decimal(row_dict.get("Net Dividend Amount"))
                qty = _to_decimal(row_dict.get("Quantity"))
                per_share = _to_decimal(row_dict.get("Dividend Per Share"))
            except BrokerFormatError:
                continue
            if net <= ZERO:
                continue
            yield NormalizedDividend(
                broker_key=self.key,
                account_label=account_label,
                isin=str(row_dict.get("ISIN") or "").strip(),
                symbol=symbol,
                ex_date=ex_date,
                pay_date=None,  # Zerodha XLSX does not report bank-credit date
                amount_gross=net,
                amount_net=net,
                tds=ZERO,
                dividend_per_share=per_share,
                quantity=qty,
                currency="INR",
                raw={"source_file": "dividends.xlsx"},
            )

    def parse_corporate_actions(
        self, file_bytes: bytes, *, account_label: str = "Main"
    ) -> Iterable[NormalizedCA]:
        # Zerodha does not publish a machine-readable corp-actions export.
        # Users enter splits/bonuses manually in Phase C.
        return iter(())

    def parse_client_id(self, file_bytes: bytes) -> str:
        """Extract the 6-char Zerodha client code from the preamble.

        The preamble row reads ``(None, 'Client ID', '<CLIENT_ID>', None, ...)``
        somewhere in the first dozen rows. We scan the active sheet and
        return the first non-empty cell that follows a ``Client ID`` label
        cell. Returns ``""`` for non-XLSX bytes -- the file's main parser
        will raise the actual error.
        """
        try:
            rows = _load_rows(file_bytes)
        except BrokerFormatError:
            return ""
        for row_idx, row in enumerate(rows):
            if row_idx > 12:
                break
            for i, cell in enumerate(row):
                if cell is None:
                    continue
                if str(cell).strip().lower() == "client id":
                    for nxt in row[i + 1 :]:
                        if nxt is None:
                            continue
                        val = str(nxt).strip()
                        if val:
                            return val.upper()
        return ""
